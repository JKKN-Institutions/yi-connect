#!/usr/bin/env python3
"""Build the circulated Yi SRTN × TN Budget 2026-27 spreadsheet.

WHY THIS EXISTS
    The public page at /tn-budget-alignment and the workbook that gets emailed
    to chapter chairs must say the same thing. They will not, if the workbook is
    maintained by hand — someone edits the page, the spreadsheet quietly goes
    stale, and two versions of "what Yi offers the Government" end up in
    circulation. So the workbook is generated FROM the page's own dataset, and
    is never edited directly.

SINGLE SOURCE OF TRUTH
    lib/policy-alignment/tn-budget-2026.ts

    That file is TypeScript, but every export is a plain JSON literal precisely
    so this script can read it without a JS runtime. If you ever add a computed
    value or a trailing comma to those arrays, this parser will stop working —
    keep them literal.

OUTPUT
    artifacts/Yi-SRTN-x-TN-Budget-2026-27-Alignment.xlsx

    artifacts/ is gitignored, so the workbook itself is deliberately NOT in the
    repo. Regenerate it whenever the page changes; do not commit the .xlsx.

USAGE
    python3 scripts/build_tn_budget_workbook.py            # from the repo data
    python3 scripts/build_tn_budget_workbook.py --ref origin/master

    Requires openpyxl (pip install openpyxl). No other dependency.

FRAMING — the reason the column order is what it is
    Every sheet leads with what the DEPARTMENT gains, not what Yi does. A Yi
    growth target is not a government benefit. If you add a column, keep that
    ordering, and read the Positioning sheet before showing any of this to an
    official.
"""

from __future__ import annotations

import argparse
import json
import re
import subprocess
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment as Al
    from openpyxl.styles import Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

REPO = Path(__file__).resolve().parent.parent
DATA_REL = "lib/policy-alignment/tn-budget-2026.ts"
OUT = REPO / "artifacts" / "Yi-SRTN-x-TN-Budget-2026-27-Alignment.xlsx"

# Palette mirrors the public page so a reader moving between them is not
# relearning colours. Green always means "what the department gets".
HEAD, INK = "17505E", "1B2A33"
GAIN, WARN, NO_, GREY = "DCEBD2", "FBEBCE", "F3E1DD", "ECEFF0"
BAND_TONE = {
    "Act now": "DCEBD2",
    "Open a door": "D6E7F2",
    "Re-scope first": "FBEBCE",
    "Not this year": "ECEFF0",
}
SCOPE_LABEL = {"chapter": "Every chapter", "region": "Once for SRTN"}

_thin = Side(style="thin", color="C6D0D4")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ── read the dataset ─────────────────────────────────────────────────────────

def load_source(ref: str | None) -> str:
    """Read the TS dataset, either from the working tree or from a git ref."""
    if ref:
        return subprocess.run(
            ["git", "show", f"{ref}:{DATA_REL}"],
            cwd=REPO, capture_output=True, text=True, check=True,
        ).stdout
    return (REPO / DATA_REL).read_text(encoding="utf-8")


def parse_array(src: str, name: str) -> list:
    m = re.search(rf"export const {name}: \w+\[\] = (\[.*?\n\]);", src, re.S)
    if not m:
        sys.exit(f"could not find export {name} — has the dataset shape changed?")
    return json.loads(m.group(1))


def parse_object(src: str, name: str) -> dict:
    m = re.search(rf"export const {name}: \w+ = (\{{.*?\n\}});", src, re.S)
    if not m:
        sys.exit(f"could not find export {name} — has the dataset shape changed?")
    return json.loads(m.group(1))


# ── sheet builder ────────────────────────────────────────────────────────────

def add_sheet(wb, name, header, rows, widths, wrap, *,
              table=None, band_col=None, bold=(), freeze="A2"):
    ws = wb.create_sheet(name)
    ws.append(header)
    for c in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=HEAD)
        cell.alignment = Al(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 42

    for row in rows:
        ws.append(row)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    for r in range(2, ws.max_row + 1):
        for c in range(1, len(header) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.font = Font(size=10, color=INK)
            cell.alignment = Al(
                vertical="top",
                wrap_text=c in wrap,
                horizontal="center" if isinstance(cell.value, (int, float)) else "left",
            )
        for c in bold:
            ws.cell(row=r, column=c).font = Font(size=10, bold=True, color=INK)
        if band_col:
            b = ws.cell(row=r, column=band_col)
            b.fill = PatternFill("solid", fgColor=BAND_TONE.get(b.value, "FFFFFF"))
            b.font = Font(size=10, bold=True, color=INK)

    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    if table and ws.max_row > 1:
        t = Table(displayName=table,
                  ref=f"A1:{get_column_letter(len(header))}{ws.max_row}")
        t.tableStyleInfo = TableStyleInfo(name="TableStyleLight1", showRowStripes=True)
        ws.add_table(t)
    return ws


def tint(ws, col, colour):
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=colour)


# ── the sheets ───────────────────────────────────────────────────────────────

README_ROWS = [
    ("Yi SRTN  ×  Tamil Nadu Budget 2026-27", ""),
    ("What the Government gains, and how Yi should position itself", ""),
    ("", ""),
    ("Author", "Ommsharravana — Chapter Chair, Yi Erode"),
    ("Region", "SRTN — 15 Yi chapters across Tamil Nadu"),
    ("Live page", "https://yi-connect-app.vercel.app/tn-budget-alignment"),
    ("Generated by", "scripts/build_tn_budget_workbook.py — do not edit this file by hand"),
    ("", ""),
    ("FRAMING", "Every row states what the DEPARTMENT gains — never what Yi gains."),
    ("", "A Yi growth target (school counts, membership) is not a government"),
    ("", "benefit and must never be presented as one. See the Positioning sheet."),
    ("", ""),
    ("Source 1", "Yi Leadership Academy 2026 — Pathfinder vertical one-pagers (National Yi)"),
    ("Source 2", "TN Budget Speech, 5 August 2026 — Revised Budget Estimates"),
    ("Citations", "The 'Budget ¶' column is the paragraph number in the speech. Every row is checkable."),
    ("", ""),
    ("SCORING", "Four dimensions, 0-2 each, total out of 8. A judgement against a fixed rubric."),
    ("Outcome", "2 identical outcome · 1 contributing · 0 unrelated"),
    ("People", "2 same beneficiary cohort · 1 overlapping · 0 different"),
    ("Timing", "2 both live in 2026-27 · 1 one side undated or next cycle · 0 not yet real"),
    ("Invitation", "2 budget text names outside participation · 1 plausible · 0 closed"),
    ("Bands", "7-8 Act now · 5-6 Open a door · 3-4 Re-scope first · 0-2 Not this year"),
    ("", ""),
    ("EFFORT", "Afternoon = one sitting · A week = a few visits · A quarter = sustained delivery"),
    ("Scope", "'Every chapter' repeats in all 15. 'Once for SRTN' is done once by the region."),
    ("Windows", "Forward-looking on YI's calendar. The State's own procurement and"),
    ("", "rollout timing is not known and is not implied anywhere in this file."),
    ("", ""),
    ("Allocation", "₹ crore, Revised Budget Estimate 2026-27, for the named scheme."),
    ("", "This is the SIZE OF THE POLICY AREA — it is NOT funding available to Yi."),
    ("", ""),
    ("CAVEAT", "Departmental constraints are inferred from the budget text and have NOT"),
    ("", "been confirmed with any department. Scale figures are Yi-side estimates,"),
    ("", "not commitments. Do not promise a department a number from this file."),
]


def build_readme(wb):
    ws = wb.create_sheet("Read Me")
    for i, (a, b) in enumerate(README_ROWS, 1):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
    ws["A1"].font = Font(bold=True, size=18, color=HEAD)
    ws["A2"].font = Font(size=12, color="43535F", italic=True)
    for r, (a, _) in enumerate(README_ROWS, 1):
        if a and r > 2:
            ws.cell(row=r, column=1).font = Font(bold=True, size=10, color=INK)
        ws.cell(row=r, column=2).font = Font(size=10, color=INK)
        ws.cell(row=r, column=2).alignment = Al(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 100
    ws.sheet_view.showGridLines = False


def build(ref: str | None) -> dict:
    src = load_source(ref)
    depts = parse_array(src, "DEPT_CASES")
    rows = parse_array(src, "ALIGNMENT_ROWS")
    chapters = parse_array(src, "CHAPTER_ROWS")
    gaps = parse_array(src, "GAP_ROWS")
    leads = parse_array(src, "LEAD_ROWS")
    pos = parse_object(src, "POSITIONING")

    # Leads carry no contact details anywhere — the page is public and the
    # Pathfinder one-pagers they came from contain personal mobile numbers.
    for l in leads:
        l.pop("contact", None)

    wb = Workbook()
    default_sheet = wb.active
    if default_sheet is not None:  # openpyxl types active as Optional
        wb.remove(default_sheet)
    wb.properties.creator = "Ommsharravana, Chapter Chair, Yi Erode"
    wb.properties.title = "Yi SRTN x Tamil Nadu Budget 2026-27 Alignment"

    build_readme(wb)

    # What Govt Gains — the centrepiece, one row per offer.
    gov = [[d["dept"], d["budget"], d["para"], d["mandate"], d["constraint"],
            o["work"], o["gains"], o["costs"], o.get("scale", ""), d["evidence"]]
           for d in depts for o in d["offers"]]
    ws = add_sheet(wb, "What Govt Gains",
        ["Department", "Budget", "Budget ¶", "Obliged to deliver",
         "What they are short of", "WHAT YI DOES", "WHAT THE DEPARTMENT GAINS",
         "What it costs them", "Scale", "Evidence handed back"],
        gov, [30, 18, 12, 56, 56, 54, 56, 34, 30, 50],
        {4, 5, 6, 7, 8, 9, 10}, table="GovtGains", bold=(1, 7), freeze="B2")
    tint(ws, 7, GAIN)
    tint(ws, 5, WARN)

    # Positioning.
    p = [["HEADLINE", pos["headline"], ""], ["", pos["sub"], ""], ["", "", ""]]
    p += [["ASSET", a["title"], a["body"]] for a in pos["assets"]] + [["", "", ""]]
    p += [["SAY", s, ""] for s in pos["say"]] + [["", "", ""]]
    p += [["NEVER SAY", s, ""] for s in pos["neverSay"]] + [["", "", ""]]
    p += [["THE ONLY ASK", s, ""] for s in pos["ask"]]
    ws = add_sheet(wb, "Positioning", ["", "Point", "Detail"], p,
                   [16, 80, 80], {2, 3}, bold=(1,))
    fills = {"SAY": GAIN, "NEVER SAY": NO_, "ASSET": "D6E7F2",
             "THE ONLY ASK": GREY, "HEADLINE": WARN}
    for r in range(2, ws.max_row + 1):
        f = fills.get(ws.cell(row=r, column=1).value)
        if f:
            ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=f)

    # Alignment matrix, highest-scoring first.
    m = []
    for r in sorted(rows, key=lambda x: (-x["score"], x["vertical"])):
        s = r["scores"]
        m.append([r["pillar"], r["vertical"], r["programme"], r["target"], r["scheme"],
                  r["para"], r["crore"], r["dept"], s["outcome"], s["people"],
                  s["timing"], s["invitation"], r["score"], r["band"], r["why"],
                  r["govGain"], r["firstStep"], r["effort"], r["ownerRole"],
                  SCOPE_LABEL.get(r["ownerScope"], ", ".join(r["namedChapters"])),
                  r["window"], r["proof"], r["action"]])
    ws = add_sheet(wb, "Alignment Matrix",
        ["Yi Pillar", "Vertical", "Yi Programme (Pathfinder 2026)",
         "Yi National Target 2026", "Tamil Nadu Scheme (Budget 2026-27)", "Budget ¶",
         "₹ cr", "TN Department", "Outcome /2", "People /2", "Timing /2",
         "Invitation /2", "SCORE /8", "Band", "Why this score",
         "WHAT THE DEPARTMENT GAINS", "DO THIS FIRST", "Effort", "Owner (role)",
         "Scope", "Window", "Proof it worked", "Strategic note"],
        m, [15, 20, 34, 32, 42, 10, 8, 24, 9, 9, 9, 10, 9, 15, 50, 54, 52, 11, 28, 17, 30, 38, 44],
        {3, 4, 5, 8, 15, 16, 17, 19, 21, 22, 23},
        table="AlignmentMatrix", band_col=14, bold=(2, 13, 16, 17), freeze="C2")
    tint(ws, 16, GAIN)
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=7).number_format = "#,##0"

    # By owner — each role's own list. "Not this year" rows have no action.
    grouped = defaultdict(list)
    for r in rows:
        if r["band"] != "Not this year":
            grouped[(r["ownerRole"], r["ownerScope"])].append(r)
    owner = []
    for (role, scope), lst in sorted(grouped.items(),
                                     key=lambda kv: (-len(kv[1]), kv[0][0])):
        for r in sorted(lst, key=lambda x: -x["score"]):
            owner.append([role, SCOPE_LABEL.get(scope, "Named chapters"), len(lst),
                          r["score"], r["programme"], r["firstStep"], r["govGain"],
                          r["effort"], r["window"], r["para"]])
    ws = add_sheet(wb, "By Owner",
        ["Owner (role)", "Scope", "Total actions", "Score /8", "Programme",
         "DO THIS FIRST", "What the department gains", "Effort", "Window", "Budget ¶"],
        owner, [30, 17, 13, 9, 34, 54, 54, 11, 30, 10], {5, 6, 7, 9},
        table="ByOwner", bold=(1, 6))
    tint(ws, 7, GAIN)

    add_sheet(wb, "Chapter Opportunities",
        ["Chapter", "District named in the Budget?", "Named scheme anchoring this city",
         "Budget ¶", "Lead vertical", "First move"],
        [[c["chapter"], c["named"], c["scheme"], c["para"], c["vertical"], c["move"]]
         for c in chapters],
        [16, 34, 50, 12, 20, 52], {2, 3, 6}, table="ChapterOpps", bold=(1,))

    add_sheet(wb, "Gaps & Do Not Chase",
        ["Type", "Item", "Detail", "Budget ¶", "₹ cr", "Recommendation",
         "DO THIS FIRST", "Effort", "Owner", "Window"],
        [[g["type"], g["item"], g["detail"], g["para"], g["crore"],
          g["recommendation"], g["firstStep"], g["effort"], g["owner"], g["window"]]
         for g in gaps],
        [22, 30, 52, 11, 9, 54, 56, 11, 28, 34], {3, 6, 7, 9, 10},
        table="GapsSheet", bold=(1, 7))

    add_sheet(wb, "SRTN Leads",
        ["Vertical", "SRTN / National Lead", "Chapter", "Role"],
        [[l["vertical"], l["name"], l["chapter"], l["role"]] for l in leads],
        [24, 30, 16, 20], set(), table="SRTNLeads", bold=(1,))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return {"depts": len(depts), "offers": len(gov), "matrix": len(m),
            "owner": len(owner), "gaps": len(gaps), "leads": len(leads),
            "bands": dict(Counter(r["band"] for r in rows)),
            "sheets": wb.sheetnames}


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--ref", help="read the dataset from a git ref (e.g. origin/master) "
                                  "instead of the working tree")
    args = ap.parse_args()

    stats = build(args.ref)
    print(f"wrote {OUT.relative_to(REPO)}")
    print(f"  sheets     : {', '.join(stats['sheets'])}")
    print(f"  departments: {stats['depts']}  offers: {stats['offers']}")
    print(f"  matrix     : {stats['matrix']} rows   by-owner: {stats['owner']}")
    print(f"  bands      : {stats['bands']}")
    print("\nartifacts/ is gitignored — do not commit the .xlsx.")


if __name__ == "__main__":
    main()
